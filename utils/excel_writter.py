from openpyxl import Workbook, load_workbook

wb = Workbook()

wb.create_sheet('cotação')
wb.create_sheet('taxas de juros')
wb.create_sheet('logs')
print(wb.sheetnames)
wb.remove(wb['Sheet'])

print(wb.sheetnames)

wb.save('desafio2.xlsx')
